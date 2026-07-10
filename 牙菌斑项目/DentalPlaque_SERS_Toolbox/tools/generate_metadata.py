#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Generate standalone PatientRegistry for clinician fill-out.  3 groups."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r"E:\牙菌斑项目\PatientRegistry_医生填写版.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "PatientRegistry"

# ── Styles ──────────────────────────────────────────────────────
DARK_BLUE = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
WHITE_BOLD = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
YELLOW     = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LIGHT_RED  = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
GRAY       = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
NORM       = Font(name="Microsoft YaHei", size=10)
RED_FONT   = Font(name="Microsoft YaHei", size=10, color="CC0000", bold=True)
TITLE_FONT = Font(name="Microsoft YaHei", size=14, bold=True, color="1F3864")
INFO_FONT  = Font(name="Microsoft YaHei", size=9, italic=True, color="666666")
THIN = Border(left=Side("thin"), right=Side("thin"),
              top=Side("thin"), bottom=Side("thin"))
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Title & info ─────────────────────────────────────────────────
ws.merge_cells("A1:K1")
ws.cell(1,1,"牙菌斑 SERS-AI 项目 — 临床样本信息登记表（医生填写版）").font = TITLE_FONT
ws.cell(1,1).alignment = CTR
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:K2")
ws.cell(2,1,"黄色单元格 = 需医生填写    红色姓名 = 姓名不完整/疑似有误    红色备注 = 需确认    仅 DoctorConfirmed=Yes 进入 AI").font = INFO_FONT
ws.cell(2,1).alignment = CTR
ws.row_dimensions[2].height = 22

# ── Header row 3 ─────────────────────────────────────────────────
HEADERS = ["患者编号 (PatientID)","患者姓名 (PatientName)","性别 (Gender)",
           "年龄 (Age)","医院 (Hospital)","当前标签","临床标签 (ClinicalLabel)",
           "医生确认 (DoctorConfirmed)","样本类型 (SampleType)",
           "采样日期 (SampleDate)","备注 (Notes)"]
N = len(HEADERS)
for c,h in enumerate(HEADERS,1):
    cell = ws.cell(3,c,h)
    cell.fill = DARK_BLUE; cell.font = WHITE_BOLD; cell.alignment = CTR; cell.border = THIN
ws.row_dimensions[3].height = 36

# ── Patients ─────────────────────────────────────────────────────
positive = [
    ("P001","顾和均"),("P002","侯长明"),("P003","黄瑜"),("P004","江学友"),
    ("P005","蒋素华"),("P006","龙有飞"),("P007","刘光晓"),("P008","刘红霞"),
    ("P009","刘盘"),("P010","刘顺利"),("P011","史成娟"),("P012","石月霞"),
    ("P013","王建江"),("P014","杨云"),("P015","羊志英"),("P016","羊志英"),
    ("P017","周士程"),("P018","张绍军"),("P019","张伟平"),("P020","章正谊"),
    ("P021","黄峰"),("P022","马玉露"),("P023","宋凡"),("P024","苏钰清"),
    ("P025","薛英"),("P026","徐贤春"),("P027","张桂平"),("P028","纪江"),
    ("P029","张翠芳"),("P030","康万里"),("P031","徐春宁"),
]
negative = [
    ("P032","卞方星"),("P033","陈惠琳"),("P034","施泽"),("P035","刘盘"),
    ("P036","刘XX？"),("P037","吕本清"),("P038","孙木林"),("P039","王超"),
    ("P040","王重阳"),("P041","王缙"),("P042","薛X刚？"),("P043","杨程雯"),
    ("P044","张弛"),("P045","张X志？"),("P046","张礼强"),("P047","张玮阳"),
    ("P048","张逸凡"),("P049","张永辰"),("P050","周子杰"),("P051","昙？亚杰"),
    ("P052","21号（XXX？）"),
]

# Unknown group: user-specified order
unknown_raw = [
    "毕家芹","蔡维超","邓X？煊","范震","管正国","姜余平",
    "孔祥宏","李广宇","刘志学","毛冬云","孔树威","陶帅文",
    "咸？桂","薛英","薛宇露","杨程雯","张怀娣",
]

unknown = []
for i, name in enumerate(unknown_raw):
    pid = f"P{53+i:03d}"
    unknown.append((pid, name))

def name_questionable(name):
    return any(ch in name for ch in ["X","？","?","XX","XXX","号（"])

def build_notes(pid, name, current_label, pid_counter):
    parts = []
    if name_questionable(name):
        parts.append("患者姓名待医生确认")
    if pid_counter.get(name, 0) > 1:
        parts.append("疑似重复采样或同名患者")
    # cross-group check
    if current_label == "Unknown" and name in set(n for _, n in positive + negative):
        parts.append("同时出现在已知组名单中，请医生确认")
    return "；".join(parts)

# Count name occurrences
all_names = [n for _, n in positive + negative + unknown]
name_count = {n: all_names.count(n) for n in all_names}

YELLOW_COLS = {3, 4, 7, 8, 10}

r = 4
all_rows = []
for pid, name in positive:
    all_rows.append((pid, name, "Positive"))
for pid, name in negative:
    all_rows.append((pid, name, "Negative"))
for pid, name in unknown:
    all_rows.append((pid, name, "Unknown"))

for pid, name, lbl in all_rows:
    notes = build_notes(pid, name, lbl, name_count)
    bad = name_questionable(name)

    ws.cell(r,1,pid);   ws.cell(r,2,name)
    ws.cell(r,3,"");    ws.cell(r,4,"")
    ws.cell(r,5,"南京大学附属鼓楼医院")
    ws.cell(r,6,lbl)
    ws.cell(r,7,"");    ws.cell(r,8,"No")
    ws.cell(r,9,"Supragingival Plaque (龈上菌斑)")
    ws.cell(r,10,"");   ws.cell(r,11,notes)

    for c in range(1,N+1):
        cell = ws.cell(r,c)
        cell.font = NORM; cell.border = THIN; cell.alignment = CTR
        if c in YELLOW_COLS:
            cell.fill = YELLOW
        # Name highlighting
        if c == 2:
            if bad:
                cell.fill = LIGHT_RED
                cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="CC0000")
        # Grey stripe for even rows (fixed columns only)
        if r % 2 == 0 and c not in YELLOW_COLS:
            existing = cell.fill
            if existing == PatternFill():
                cell.fill = GRAY

    if notes:
        ws.cell(r,11).font = RED_FONT

    # Group separator: thin darker border between groups
    if pid == "P032":
        for c2 in range(1,N+1):
            ws.cell(r,c2).border = Border(
                top=Side("medium"), left=Side("thin"),
                right=Side("thin"), bottom=Side("thin"))
    if pid == "P053":
        for c2 in range(1,N+1):
            ws.cell(r,c2).border = Border(
                top=Side("medium"), left=Side("thin"),
                right=Side("thin"), bottom=Side("thin"))

    ws.row_dimensions[r].height = 22
    r += 1

LAST = r - 1  # P069

# ── Dropdowns ────────────────────────────────────────────────────
for dv_formula, col_letter in [
    ('"Male,Female,Unknown"', "C"),
    ('"Healthy,Gingivitis,Periodontitis"', "G"),
    ('"Yes,No"', "H"),
]:
    dv = DataValidation(type="list", formula1=dv_formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}4:{col_letter}{LAST}")

# ── Freeze / filter / widths ─────────────────────────────────────
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:K{LAST}"
for c, w in {1:11,2:14,3:9,4:8,5:26,6:14,7:14,8:12,9:28,10:13,11:42}.items():
    ws.column_dimensions[get_column_letter(c)].width = w
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
ws.print_title_rows = "3:3"

wb.save(OUT)

nPos = len(positive); nNeg = len(negative); nUnk = len(unknown)
print(f"Saved: {OUT}")
print(f"  Positive: {nPos} patients (P001-P{1+nPos-1:03d})")
print(f"  Negative: {nNeg} patients (P032-P{32+nNeg-1:03d})")
print(f"  Unknown:  {nUnk} patients (P053-P{53+nUnk-1:03d})")
print(f"  Total:    {nPos+nNeg+nUnk}")
print(f"  Red names: {sum(1 for _, n, _ in all_rows if name_questionable(n))}")
